import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
import datetime

def create_tracker():
    wb = openpyxl.Workbook()
    
    # ─── 1. SETUP SHEET: PROJECT PLAN ──────────────────────────────────────────
    ws_track = wb.active
    ws_track.title = "Project Plan"
    
    # Enable grid lines
    ws_track.views.sheetView[0].showGridLines = True
    
    # Data from the user's screenshots
    tracker_data = [
        {
            'Model ID': 'M1',
            'Module': 'Login',
            'Model Name': 'Authentication & Authorization',
            'Sub Model': 'Authentication',
            'Description': 'Implement secure Email/Mobile and Password login with validation, JWT authentication, error handling, password visibility toggle, and password strength indicator.',
            'Open Date': None,
            'Close Date': None,
            'Development Status': 'Completed',
            'Development Time': '15 days',
            'Test Case ID': (
                'TC-1.0-01: Initial UI Load;\n'
                'TC-1.0-02: Validate empty email/mobile input;\n'
                'TC-1.0-03: Validate mobile number format;\n'
                'TC-1.0-04: Validate email format;\n'
                'TC-1.0-05: Detect mobile input;\n'
                'TC-1.0-06: Password visibility toggle;\n'
                'TC-1.0-07: Password strength indicator;'
            )
        },
        {
            'Model ID': 'M2',
            'Module': 'OrgRolePage',
            'Model Name': 'Authentication & Authorization',
            'Sub Model': 'Role Selection',
            'Description': 'Dynamically load workspaces from API, allow workspace selection, and display user role based on the selected workspace.',
            'Open Date': None,
            'Close Date': None,
            'Development Status': 'Completed',
            'Development Time': None,
            'Test Case ID': (
                'TC-1.1-01: Warehouse Dropdown Load;\n'
                'TC-1.1-02: Warehouse Dropdown Select;\n'
                'TC-1.1-03: Warehouse Filtering;\n'
                'TC-1.1-04: Workspace Change;'
            )
        },
        {
            'Model ID': 'M2',
            'Module': '',  # Blank as in screenshot
            'Model Name': 'Authentication & Authorization',
            'Sub Model': 'Role Metrics',
            'Description': 'Dynamically filter and display operational metrics based on the selected workspace and role using dashboard Stat Cards.',
            'Open Date': None,
            'Close Date': None,
            'Development Status': 'Completed',
            'Development Time': None,
            'Test Case ID': (
                'TC-1.2-01: Role Dropdown Load;\n'
                'TC-1.2-02: Role Filtering;\n'
                'TC-1.2-03: Display selected role metrics;'
            )
        }
    ]
    
    headers = [
        "Model ID", "Module", "Model Name", "Sub Model", "Description", 
        "Open Date", "Close Date", "Development Status", "Development Time", "Test Case ID"
    ]
    
    # Write Headers
    for c_idx, h in enumerate(headers, 1):
        ws_track.cell(row=1, column=c_idx, value=h)
        
    # Write Rows
    for r_idx, row_data in enumerate(tracker_data, 2):
        ws_track.cell(row=r_idx, column=1, value=row_data['Model ID'])
        ws_track.cell(row=r_idx, column=2, value=row_data['Module'])
        ws_track.cell(row=r_idx, column=3, value=row_data['Model Name'])
        ws_track.cell(row=r_idx, column=4, value=row_data['Sub Model'])
        ws_track.cell(row=r_idx, column=5, value=row_data['Description'])
        
        # Open Date
        open_cell = ws_track.cell(row=r_idx, column=6, value=row_data['Open Date'])
        open_cell.number_format = 'yyyy-mm-dd'
        
        # Close Date
        close_cell = ws_track.cell(row=r_idx, column=7, value=row_data['Close Date'])
        close_cell.number_format = 'yyyy-mm-dd'
        
        # Development Status
        ws_track.cell(row=r_idx, column=8, value=row_data['Development Status'])
        
        # Development Time: If value is provided use it, otherwise write formula if dates are filled
        if row_data['Development Time'] is not None:
            ws_track.cell(row=r_idx, column=9, value=row_data['Development Time'])
        else:
            time_formula = f'=IF(AND(F{r_idx}<>"", G{r_idx}<>""), G{r_idx}-F{r_idx}, "")'
            time_cell = ws_track.cell(row=r_idx, column=9, value=time_formula)
            time_cell.number_format = '0" days"'
        
        # Test Case ID
        ws_track.cell(row=r_idx, column=10, value=row_data['Test Case ID'])

    # Styling Elements (Arial Font to match Google Sheets exactly, sizes 10 & 11)
    # Header styling: White background, bold black text, centered vertically, left horizontally
    header_font = Font(name="Arial", size=10, bold=True, color="000000")
    header_align = Alignment(horizontal="left", vertical="center")
    
    border_side = Side(style='thin', color='D9D9D9')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Format Header Row
    ws_track.row_dimensions[1].height = 28
    for col in range(1, 11):
        cell = ws_track.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border
        
    # Format Data Rows
    for r_idx in range(2, len(tracker_data) + 2):
        # Allow row height to auto-expand for wrapped text, do not specify hardcoded height!
        ws_track.row_dimensions[r_idx].height = None
        
        for c_idx in range(1, 11):
            cell = ws_track.cell(row=r_idx, column=c_idx)
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.border = cell_border
            
            # Alignments (Left & Top as per Google Sheets default)
            if c_idx in [5, 10]:  # Description and Test Cases
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top")
                
    # Freeze Header Row (row 1 is frozen, scrollable below A2)
    ws_track.freeze_panes = 'A2'
    
    # Add Filter to Header
    ws_track.auto_filter.ref = f"A1:J{len(tracker_data) + 1}"
    
    # Dropdown Validation for Development Status (Column H)
    dv = DataValidation(type="list", formula1='"Not Started,Progress,In Progress,Completed,Blocked,UAT,Deployed"', allow_blank=True)
    dv.error = 'Your entry is not in the list of allowed statuses'
    dv.errorTitle = 'Invalid Status'
    dv.prompt = 'Select development status'
    dv.promptTitle = 'Status'
    ws_track.add_data_validation(dv)
    dv.add(f"H2:H100")  # Apply to future rows as well
    
    # Conditional Formatting Rules for Status Column (Column H)
    # Nice professional pastel status colors
    status_cf_colors = {
        "Completed": {"fill": "E2EFDA", "font": "375623"},      # Green
        "In Progress": {"fill": "DDEBF7", "font": "1F4E79"},    # Blue
        "Progress": {"fill": "E8F1F5", "font": "2F5597"},       # Light Blue
        "Blocked": {"fill": "FCE4D6", "font": "C00000"},        # Red
        "Not Started": {"fill": "F2F2F2", "font": "595959"},    # Gray
        "UAT": {"fill": "FFF2CC", "font": "7F6000"},            # Orange
        "Deployed": {"fill": "C6EFCE", "font": "006100"}        # Dark Green
    }
    
    for val, colors in status_cf_colors.items():
        cf_fill = PatternFill(start_color=colors["fill"], end_color=colors["fill"], fill_type="solid")
        cf_font = Font(name="Arial", size=10, bold=True, color=colors["font"])
        rule = CellIsRule(operator='equal', formula=[f'"{val}"'], fill=cf_fill, font=cf_font)
        ws_track.conditional_formatting.add(f"H2:H100", rule)
        
    # Highlight Overdue Items Formula Rule (applied to entire row range A2:J100)
    overdue_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    overdue_font = Font(name="Arial", size=10, color="9C0006", bold=True)
    overdue_rule = FormulaRule(
        formula=[f'AND($H2<>"Completed", $H2<>"Deployed", $F2<>"", TODAY()-$F2>30)'],
        fill=overdue_fill,
        font=overdue_font
    )
    ws_track.conditional_formatting.add(f"A2:J100", overdue_rule)

    # Set precise column widths to match screenshots perfectly
    column_widths = {
        'A': 12,  # Model ID
        'B': 18,  # Module
        'C': 32,  # Model Name
        'D': 22,  # Sub Model
        'E': 65,  # Description
        'F': 14,  # Open Date
        'G': 14,  # Close Date
        'H': 22,  # Development Status
        'I': 20,  # Development Time
        'J': 60,  # Test Case ID
    }
    for col_letter, width in column_widths.items():
        ws_track.column_dimensions[col_letter].width = width


    # ─── 2. SETUP SHEET: SUMMARY DASHBOARD ─────────────────────────────────────
    ws_dash = wb.create_sheet(title="Summary Dashboard", index=1)
    ws_dash.views.sheetView[0].showGridLines = True
    
    dash_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    dash_header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    
    # Merged title block
    ws_dash.merge_cells("B2:L2")
    title_cell = ws_dash["B2"]
    title_cell.value = "Project Development & QA Dashboard"
    title_cell.fill = dash_header_fill
    title_cell.font = dash_header_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[2].height = 40
    
    # Key Metrics Cards Block
    metrics_labels = [
        ("Total Modules", "=SUMPRODUCT(('Project Plan'!A$2:A$100<>\"\")/COUNTIF('Project Plan'!A$2:A$100, 'Project Plan'!A$2:A$100&\"\"))"),
        ("Total Sub Models", "=COUNTIF('Project Plan'!D$2:D$100, \"?*\")"),
        ("Overall Completion %", "=(COUNTIF('Project Plan'!H$2:H$100, \"Completed\") + COUNTIF('Project Plan'!H$2:H$100, \"Deployed\"))/C6")
    ]
    
    # Write Key Metrics Table
    ws_dash.cell(row=4, column=2, value="Key Metrics").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=4, column=3, value="Value").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=4, column=2).border = Border(bottom=Side(style='medium', color='1F4E79'))
    ws_dash.cell(row=4, column=3).border = Border(bottom=Side(style='medium', color='1F4E79'))
    
    for idx, (label, formula) in enumerate(metrics_labels, 5):
        lbl_cell = ws_dash.cell(row=idx, column=2, value=label)
        lbl_cell.font = Font(name="Arial", size=10, bold=True)
        lbl_cell.border = cell_border
        
        val_cell = ws_dash.cell(row=idx, column=3, value=formula)
        val_cell.font = Font(name="Arial", size=10, bold=True, color="1F4E79")
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = cell_border
        
        if label == "Overall Completion %":
            val_cell.number_format = '0.0%'
        else:
            val_cell.number_format = '0'
            
    # Write Status Breakdown Table
    ws_dash.cell(row=9, column=2, value="Status Breakdown").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=9, column=3, value="Count").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=9, column=2).border = Border(bottom=Side(style='medium', color='1F4E79'))
    ws_dash.cell(row=9, column=3).border = Border(bottom=Side(style='medium', color='1F4E79'))
    
    statuses = ["Not Started", "Progress", "In Progress", "Completed", "Blocked", "UAT", "Deployed"]
    for idx, status in enumerate(statuses, 10): # Row 10 to 16
        lbl_cell = ws_dash.cell(row=idx, column=2, value=status)
        lbl_cell.font = Font(name="Arial", size=10)
        lbl_cell.border = cell_border
        
        val_cell = ws_dash.cell(row=idx, column=3, value=f'=COUNTIF(\'Project Plan\'!H$2:H$100, "{status}")')
        val_cell.font = Font(name="Arial", size=10, bold=True)
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = cell_border
        val_cell.number_format = '0'

    # Write Module-wise Progress Table
    ws_dash.cell(row=19, column=2, value="Module").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=19, column=3, value="Sub Models").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=19, column=4, value="Completed").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws_dash.cell(row=19, column=5, value="Progress %").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    
    for c_idx in range(2, 6):
        ws_dash.cell(row=19, column=c_idx).border = Border(bottom=Side(style='medium', color='1F4E79'))
        
    modules = [
        ("Login", "M1"),
        ("OrgRolePage", "M2")
    ]
    
    for idx, (mod_name, mod_id) in enumerate(modules, 20): # row 20 to 21
        # Module Name
        lbl_cell = ws_dash.cell(row=idx, column=2, value=mod_name)
        lbl_cell.font = Font(name="Arial", size=9)
        lbl_cell.border = cell_border
        
        # Sub Models Count based on Model ID (ignores blanks/layout differences in Col B)
        total_cell = ws_dash.cell(row=idx, column=3, value=f'=COUNTIF(\'Project Plan\'!A$2:A$100, "{mod_id}")')
        total_cell.font = Font(name="Arial", size=9)
        total_cell.alignment = Alignment(horizontal="right")
        total_cell.border = cell_border
        
        # Completed & Deployed Count
        comp_cell = ws_dash.cell(row=idx, column=4, value=f'=COUNTIFS(\'Project Plan\'!A$2:A$100, "{mod_id}", \'Project Plan\'!H$2:H$100, "Completed") + COUNTIFS(\'Project Plan\'!A$2:A$100, "{mod_id}", \'Project Plan\'!H$2:H$100, "Deployed")')
        comp_cell.font = Font(name="Arial", size=9)
        comp_cell.alignment = Alignment(horizontal="right")
        comp_cell.border = cell_border
        
        # Progress %
        prog_cell = ws_dash.cell(row=idx, column=5, value=f'=IF(C{idx}>0, D{idx}/C{idx}, 0)')
        prog_cell.font = Font(name="Arial", size=9, bold=True, color="1F4E79")
        prog_cell.alignment = Alignment(horizontal="right")
        prog_cell.border = cell_border
        prog_cell.number_format = '0.0%'

    # Set Column widths for Dashboard
    ws_dash.column_dimensions['A'].width = 3
    ws_dash.column_dimensions['B'].width = 28
    ws_dash.column_dimensions['C'].width = 15
    ws_dash.column_dimensions['D'].width = 15
    ws_dash.column_dimensions['E'].width = 15
    ws_dash.column_dimensions['F'].width = 3
    
    # ─── 3. ADD CHARTS TO DASHBOARD ───────────────────────────────────────────
    # Status Pie Chart (placed at G4)
    pie = PieChart()
    pie.title = "Status Distribution"
    labels_ref = Reference(ws_dash, min_col=2, min_row=10, max_row=16)
    data_ref = Reference(ws_dash, min_col=3, min_row=9, max_row=16)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels_ref)
    pie.width = 16
    pie.height = 10
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = True
    ws_dash.add_chart(pie, "G4")
    
    # Module Column Chart (placed at G20)
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Module-wise Progress"
    bar.y_axis.title = "Progress (%)"
    bar.x_axis.title = "Module"
    
    bar_data_ref = Reference(ws_dash, min_col=5, min_row=19, max_row=21)
    bar_cats_ref = Reference(ws_dash, min_col=2, min_row=20, max_row=21)
    bar.add_data(bar_data_ref, titles_from_data=True)
    bar.set_categories(bar_cats_ref)
    bar.legend = None
    bar.width = 18
    bar.height = 10
    ws_dash.add_chart(bar, "G16") # slightly higher to look compact
    
    # Save Workbook
    wb.save("Project_Tracker.xlsx")
    print("Workbook successfully created and saved as Project_Tracker.xlsx")

if __name__ == "__main__":
    create_tracker()
